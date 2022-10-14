#To connect the SQL server and excute the sql query used pymysql lib
import pymysql
# to load the JSON file mysql_querry, which while read the database_server, User name, Password, Database name and data base port(3306) from mysql_querry file
import json
#To read and write the contain in the Excel file used openpyxl lib.
from openpyxl import load_workbook

def SelSqlFun():
    """" This function will fetch the information from table for certain critiria() """
    
    # used with clause to read the mysql_querry file and load the database_server, User name, Password, Database name and data base port in data dict object
    with open(r"mysql_querry", "r") as file:
        data = json.load(file)
    
    #used pymysql.connect method to connect the database using database_server, User name, Password, Database name and database port
    db = pymysql.connect(host=data["DB_HOST"], user=data["DB_USER"], password=data["DB_PASSWD"], database=data["DB_NAME"], port=data["DB_PORT"])
    cursor = db.cursor()
    
    # Build the sql query to fetch the details from database 
    sql = f'SELECT {data["DB_FIELD"]} FROM {data["TABLE_NAME"]} WHERE app_code IN ({data["APP_CODE"]});'
    try:
       # Execute the SQL command
       cursor.execute(sql)
        
       # Fetch all the rows in a list of lists.
       results = cursor.fetchall()
       
       #Logic to fill the template using different type of servers, different type of Operating system or Database serevr
       #Column 6(row[6]) for Type of server like PROD, DR, TEST, DEV 
       # Column 3 (row[3]) for operating system type or database server type.
       # depending upon the above value resplative template will filled. Like for PROD server and Linux operating system template1_pord_OS.xlsx will filled.
       for row in results:
           if ( row[6] == 'PROD' or row[6] == 'DR'):
               env="prod"
               if( row[3] == "Windows" or row[3] == "Linux"):
                   infra="OS"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'F{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
               elif( row[3] == "Oracle" or row[3] == "SQLServer"):
                   infra="DB"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'G{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
           elif ( row[6] == 'DEV' or row[6] == 'QA'):
               env="nonprod"
               if( row[3] == "Windows" or row[3] == "Linux"):
                   infra="OS"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'D{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
               elif( row[3] == "Oracle" or row[3] == "SQLServer"):
                   infra="DB"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'C{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
    except Exception as e:
       print (f"Error: unable to fetch data with error {e}")

    # disconnect from server
    db.close()
